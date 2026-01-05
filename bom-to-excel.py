import json
import pandas as pd
import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

def classify_category(group, name):
    """프런트엔드 라이브러리 특성에 따른 카테고리 분류 로직"""
    g = str(group).lower() if group else ""
    n = str(name).lower() if name else ""

    if any(x in n for x in ['react', 'zustand', 'router', 'context']): return 'Framework (React/Core)'
    if any(x in n for x in ['tiptap', 'prosemirror']): return 'Editor (RichText/Tiptap)'
    if any(x in n for x in ['vite', 'eslint', 'prettier', 'postcss']): return 'Development (Build/Lint)'
    if any(x in n for x in ['testing', 'vitest', 'jest', 'jsdom']): return 'Development (Testing)'
    if any(x in n for x in ['axios', 'fetch']): return 'Network (API/Http)'
    if 'types' in n or 'types' in g: return 'Development (Types/TS)'
    if any(x in n for x in ['svar-ui', 'grid', 'table']): return 'UI Component (Grid/Table)'

    return 'Other Libraries'

def run_conversion(input_path, output_path):
    if not os.path.exists(input_path):
        print(f"❌ 오류: {input_path} 파일이 없습니다.")
        return

    try:
        # 1. JSON 데이터 읽기
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # 2. 데이터 가공
        rows = []
        for comp in data.get('components', []):
            licenses = comp.get('licenses', [])
            license_str = "Unknown"
            if licenses:
                license_list = []
                for l in licenses:
                    lic = l.get('license', {})
                    name = lic.get('id') or lic.get('name') or l.get('expression')
                    if name: license_list.append(name)
                if license_list: license_str = ", ".join(license_list)

            rows.append({
                'Category': classify_category(comp.get('group'), comp.get('name')),
                'Package Name': comp.get('name'),
                'Installed Version': comp.get('version'),
                'License': license_str,
                'Description': comp.get('description') or comp.get('name')
            })

        # 3. 데이터프레임 생성 및 정렬
        df = pd.DataFrame(rows)
        df = df.sort_values(by=['Category', 'Package Name'])

        # 4. 엑셀 생성 및 스타일링
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='SBOM')
            ws = writer.sheets['SBOM']

            # 헤더 고정
            ws.freeze_panes = 'A2'

            # 스타일 정의
            header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # 헤더 및 데이터 셀 스타일/병합 적용
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = align_center
                cell.border = border

            current_cat, start_row = "", 2
            max_row = ws.max_row
            for i in range(2, max_row + 1):
                for col in range(1, 6):
                    c = ws.cell(row=i, column=col)
                    c.border = border
                    c.alignment = align_center if col <= 4 else align_left

                cat_val = ws.cell(row=i, column=1).value
                if cat_val != current_cat:
                    if i > start_row + 1:
                        ws.merge_cells(start_row=start_row, start_column=1, end_row=i-1, end_column=1)
                    current_cat, start_row = cat_val, i

            if max_row >= start_row and max_row > 2:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=max_row, end_column=1)

            # 너비 설정
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 60

        print(f"✅ 성공: {output_path} 생성 완료!")

    except Exception as e:
        print(f"❌ 오류 발생: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Convert CycloneDX bom.json to Styled Excel')
    parser.add_argument('--input', '-i', default='bom.json', help='Input JSON file path (default: bom.json)')
    parser.add_argument('--output', '-o', default='SMEP_AFE_Architecture_BOM.xlsx', help='Output Excel file path')

    args = parser.parse_args()
    run_conversion(args.input, args.output)